"""Lecture d'un export de statistiques LinkedIn.

LinkedIn ne donne rien par API pour les publications d'un membre, mais son
interface propose un export : c'est la seule source qui contienne les
impressions par post. Ce module lit ce fichier et range ce qu'il contient.

Les noms de feuilles sont ceux de l'export français et se retrouvent tronqués
(« DONNÉES DÉMOGRAPHIQUES SUR L'AU ») : on les repère donc par leur début,
jamais par égalité stricte.

Le rapprochement avec nos publications se fait sur l'identifiant numérique, que
l'URL de l'export porte en suffixe — « ...-share-7502311042279006209-fUl- » —
et que nous stockons sous la forme « urn:li:share:7502311042279006209 ».
"""

import logging
import re
from datetime import datetime

log = logging.getLogger(__name__)

_ID_RE = re.compile(r'(?:share|ugcPost|activity)[-:](\d{6,})', re.I)


class ImportError_(RuntimeError):
    pass


def _feuille(wb, debut):
    for nom in wb.sheetnames:
        if nom.strip().upper().startswith(debut.upper()):
            return wb[nom]
    return None


def _date(valeur):
    if isinstance(valeur, datetime):
        return valeur.date()
    if hasattr(valeur, 'year') and not isinstance(valeur, str):
        return valeur
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%m/%d/%Y'):
        try:
            return datetime.strptime(str(valeur).strip(), fmt).date()
        except (ValueError, TypeError):
            continue
    return None


def _entier(valeur):
    if valeur is None or valeur == '':
        return None
    try:
        return int(float(str(valeur).replace(' ', '').replace(',', '.')))
    except (TypeError, ValueError):
        return None


def post_id(url):
    """The numeric id inside a post URL, or None."""
    m = _ID_RE.search(url or '')
    return m.group(1) if m else None


def lire(fileobj):
    """Parse the export. Returns {'jours': [...], 'posts': {...}, 'abonnes': n}.

    `jours` holds one entry per day with impressions, interactions and new
    followers; `posts` maps a numeric post id to its impressions and
    interactions. Nothing is written here — the caller decides.
    """
    try:
        import openpyxl
    except ImportError as exc:                       # pragma: no cover
        raise ImportError_("openpyxl n'est pas installé sur le serveur.") from exc

    try:
        wb = openpyxl.load_workbook(fileobj, data_only=True)
    except Exception as exc:
        raise ImportError_(f"Fichier illisible : {exc}") from exc

    jours = {}

    # ── Engagement : impressions et interactions, jour par jour ───────────
    ws = _feuille(wb, 'ENGAGEMENT')
    if ws is not None:
        for ligne in ws.iter_rows(min_row=2, values_only=True):
            jour = _date(ligne[0] if ligne else None)
            if jour is None:
                continue
            entree = jours.setdefault(jour, {})
            entree['impressions'] = _entier(ligne[1] if len(ligne) > 1 else None)
            entree['interactions'] = _entier(ligne[2] if len(ligne) > 2 else None)

    # ── Abonnés : le total, puis les gains quotidiens ─────────────────────
    abonnes = None
    ws = _feuille(wb, 'ABONNÉS') or _feuille(wb, 'ABONNES') or _feuille(wb, 'FOLLOWERS')
    if ws is not None:
        for ligne in ws.iter_rows(values_only=True):
            if not ligne or ligne[0] is None:
                continue
            tete = str(ligne[0])
            # « Nombre total d'abonnés le 06/09/2026 » : le total est à côté.
            if 'total' in tete.lower():
                abonnes = _entier(ligne[1] if len(ligne) > 1 else None)
                continue
            jour = _date(tete)
            if jour is not None:
                jours.setdefault(jour, {})['nouveaux_abonnes'] = _entier(
                    ligne[1] if len(ligne) > 1 else None)

    # ── Meilleurs posts : deux tableaux côte à côte, interactions à gauche,
    #    impressions à droite, chacun avec ses propres lignes.
    posts = {}
    ws = _feuille(wb, 'MEILLEURS POSTS') or _feuille(wb, 'TOP POSTS')
    if ws is not None:
        for ligne in ws.iter_rows(values_only=True):
            if not ligne:
                continue
            # Colonnes 0-2 : interactions. Colonnes 4-6 : impressions.
            for depart, champ in ((0, 'interactions'), (4, 'impressions')):
                if len(ligne) <= depart + 2:
                    continue
                pid = post_id(str(ligne[depart] or ''))
                valeur = _entier(ligne[depart + 2])
                if pid and valeur is not None:
                    entree = posts.setdefault(pid, {})
                    entree[champ] = valeur
                    jour = _date(ligne[depart + 1])
                    if jour:
                        entree['jour'] = jour

    return {'jours': [dict(jour=j, **v) for j, v in sorted(jours.items())],
            'posts': posts, 'abonnes': abonnes}
